import openpyxl as ox
from pyswip import Prolog

course_map = [[], []]
type_map = [[], [], []]
l_map = [[], [], []]
t_map = [[], [], []]


def get_course_number(course_code):
    if not course_map[1]:
        course_map[0].append(course_code)
        course_map[1].append(1)
        return course_map[1][-1]

    elif course_code not in course_map[0]:
        course_map[0].append(course_code)
        course_map[1].append(course_map[1][-1] + 1)
        return course_map[1][-1]
    else:
        return course_map[1][course_map[0].index(course_code)]


def get_lec_group(c_number, group_name):
    if not l_map[0] or c_number not in l_map[0]:
        l_map[0].append(c_number)
        l_map[1].append(group_name)
        l_map[2].append(1)
        return 1

    if c_number in l_map[0]:
        first_i = l_map[0].index(c_number)
        end_i = len(l_map[0]) - l_map[0][::-1].index(c_number)
        if group_name not in l_map[1][first_i:end_i]:
            l_map[0].append(c_number)
            l_map[1].append(group_name)
            l_map[2].append(l_map[2][-1]+1)
            return l_map[2][-1]

    i = 0
    while i < len(l_map[0]):
        if l_map[0][i] == c_number and l_map[1][i] == group_name:
            return l_map[2][i]
        i += 1


def get_tut_group(c_number, group_name):
    if not t_map[0] or c_number not in t_map[0]:
        t_map[0].append(c_number)
        t_map[1].append(group_name)
        t_map[2].append(1)
        return 1

    if c_number in t_map[0]:
        first_i = t_map[0].index(c_number)
        end_i = len(t_map[0]) - t_map[0][::-1].index(c_number)
        if group_name not in t_map[1][first_i:end_i]:
            t_map[0].append(c_number)
            t_map[1].append(group_name)
            t_map[2].append(t_map[2][-1]+1)
            return t_map[2][-1]

    i = 0
    while i < len(t_map[0]):
        if t_map[0][i] == c_number and t_map[1][i] == group_name:
            return t_map[2][i]
        i += 1


def add_type(course_num, type_name, type_num):
    if course_num not in type_map[0]:
        type_map[0].append(course_num)
        type_map[1].append(type_name)
        type_map[2].append(type_num)
        return

    i = type_map[0].index(course_num)
    j = len(type_map[0]) - type_map[0][::-1].index(course_num) + 1

    if type_name not in type_map[1][i:j]:
        type_map[0].append(course_num)
        type_map[1].append(type_name)
        type_map[2].append(type_num)
        return


def parse_excel(workbook):
    # Initial variable values
    ws = workbook.worksheets[0]
    index = 'A2'
    row = 2
    sched = "Schedule = ["
    c_num_old = get_course_number(ws[index].value)
    old_type_txt = ws['E'+str(row)].value
    old_grp_txt = ws['G'+str(row)].value
    base_type_int = 1
    cur_type_int = 1

    # main parsing loop
    while ws[index].value:
        c_num_cur = get_course_number(ws[index].value)
        cur_type_txt = ws['E'+str(row)].value
        cur_grp_txt = ws['G'+str(row)].value

        # New entry value (Semester and Course Num)
        entry = str([int(s) for s in ws[index].value.split() if s.isdigit()][0] / 100)
        if c_num_cur < 10:
            entry += '0' + str(c_num_cur)
        else:
            entry += str(c_num_cur)

        # Determine Type value
        if c_num_cur != c_num_old:
            base_type_int = 1
            cur_type_int = 1
        elif cur_type_txt != old_type_txt:
            base_type_int = cur_type_int
        elif cur_grp_txt != old_grp_txt:
            cur_type_int = base_type_int
        add_type(c_num_cur, cur_type_txt, cur_type_int)

        # Adding Type value to entry in two digits
        if cur_type_int < 10:
            entry += '0' + str(cur_type_int)
        else:
            entry += str(cur_type_int)

        # Finding and adding Lecture group Number
        if cur_type_txt == 'Lecture':
            lec_group = get_lec_group(c_num_cur, cur_grp_txt)
            if lec_group < 10:
                entry += '0' + str(lec_group) + '00'
            else:
                entry += str(lec_group) + '00'

        # Finding and adding Tut and Lab group Number
        elif cur_type_txt == 'Tut' or cur_type_txt == 'Lab':
            tut_group = get_tut_group(c_num_cur, cur_grp_txt)
            if tut_group < 10:
                entry += '000' + str(tut_group)
            else:
                entry += '00' + str(tut_group)

        # Adding Timestamp at the end of entry
        entry += str(ws['C' + str(row)].value) + str(ws['D' + str(row)].value)

        # Concatenating to Schedule String
        sched += entry + ", "

        c_num_old = c_num_cur
        old_type_txt = cur_type_txt
        old_grp_txt = cur_grp_txt
        cur_type_int += 1
        row += 1
        index = 'A' + str(row)

    sched = sched[:-2] + "]"
    return sched


def parse_curr(curr):
    curr = [x.strip() for x in curr.split(';')]
    w_curr = []
    for course in curr:
        w_curr.append([x.strip() for x in course.split(',')])

    for c_index, course in enumerate(w_curr):
        for e_index, element in enumerate(course):
            if element.isdigit():
                course[e_index+1:] = [course[e_index+1:]]
                break

    for c_index, course in enumerate(w_curr):
        for e_index, element in enumerate(course):
            if type(element) is list:
                for ce_index, in_element in enumerate(element):
                    element[ce_index] = get_course_number(in_element)
            elif element.isdigit():
                course[e_index] = int(element)
            else:
                w_curr[c_index][e_index] = get_course_number(element)
    return w_curr


def parse_history(history):
    history = [[y.strip() for y in x.strip().split(',')] for x in history.split(';')]
    for course in history:
        course[0] = get_course_number(course[0])
    return history


def parse_oblig(oblig):
    oblig = [[y.strip() for y in x.strip().split(',')] for x in oblig.split(';')]
    oblig_courses = []
    oblig_ch = []
    for course in oblig:
        oblig_courses.append(get_course_number(course[0]))
        oblig_ch.append(int(course[1]))
    return oblig_courses, oblig_ch

# wb = ox.load_workbook("data.xlsx")
# print parse_excel(wb)
# print course_map
# print "L-map" + str(l_map)
# curr = "MATH 103, 8, CSEN 102, CSEN 501; CSEN 906, 6, CSEN 501; DE 404, 4, DE 202"
# parse_curr(curr)
# history = "MATH 103, A-; CSEN 906, B+; DE 404, FA; DMET 904, C"
# print parse_history(history)
# oblig = "MATH 103, 5; CSEN 906, 8; DE 404, 4;DMET 904, 6"
# print parse_oblig(oblig)


def get_course_name(course_num):
    return course_map[0][course_map[1].index(course_num)]


def get_type(course_num, type_num):
    i = type_map[0].index(course_num)
    j = len(type_map[0]) - type_map[0][::-1].index(course_num) + 1
    k = type_map[2][i:j].index(type_num)
    return type_map[1][k]


def get_lec_name(course_num, lec_num):
    i = l_map[0].index(course_num)
    j = len(l_map[0]) - l_map[0][::-1].index(course_num) + 1
    k = l_map[2][i:j].index(lec_num)
    return l_map[1][k]


def get_tut_name(course_num, tut_num):
    i = t_map[0].index(course_num)
    j = len(t_map[0]) - t_map[0][::-1].index(course_num) + 1
    k = t_map[2][i:j].index(tut_num)
    return t_map[1][k]


def parse_out(out):
    schedule = []
    for element in out:
        element = str(element)
        session = [get_course_name(int(element[-10:-8])), get_type(int(element[-10:-8]), int(element[-8:-6]))]

        if element[-6:-4] == '00':
            session.append(get_tut_name(int(element[-10:-8]), int(element[-4:-2])))
        else:
            session.append(get_lec_name(int(element[-10:-8]), int(element[-6:-4])))

        session.append(int(element[-2]))
        session.append(int(element[-1]))

        schedule.append(session)

    return schedule

# 60603010245, 60602010143, 60601010041, 60503010133, 60502010132, 60501010031, 60203010123, 60202010122,
#        60201010021, 60103010113, 60102010112,

# out = [60101000111]
# print parse_out(out)


# Some Extra Parsing
def final_parse(Curriculum, History, Obligatory, Schedule, Credits, Probation):
    query = ""
    query += "Oblig="
    query += str(Obligatory[0])
    query += ", ObligCH="
    query += str(Obligatory[1])
    query += ", Curr="
    query += str(Curriculum)
    query += ", History="
    query += str(History)
    query += ", Credits ="
    query += str(Credits)
    query += ", Prob="
    query += str(Probation).lower()
    query += ", "
    query += str(Schedule)
    query += ", schedule(Curr, History, Oblig, ObligCH, Credits, Schedule, Prob, 5, Solution)."
    return query

def run_prolog(query):
    prolog = Prolog()
    prolog.consult('WebApp/Project.pl')
    for result in prolog.query(query, maxresult=1):
        res = result["Solution"]
        return res